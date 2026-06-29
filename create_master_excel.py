import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Timesheet'

# Headers
headers = [
    'Project Name - L1', 'Sub Project - L2', 'Task Type', 'Task Description',
    'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'JIRA ID / SR'
]
ws.append(headers)

# Styling headers
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Sample data from screenshot
data = [
    ['MEL', 'BAU enhancement-CR', 'Meetings', 'Daily scrum', 0, 0.75, 0.75, 0.75, 0.75, 0.75, 0, 'PROJ-001'],
    ['MEL', 'BAU enhancement-CR', 'development and configuration', 'Workred on the Critical issue fix on Apis for sonarqube', 0, 4, 0, 0, 0, 0, 0, 'PROJ-002'],
    ['MEL', 'BAU enhancement-CR', 'Review', 'frontend Dashboard code merege , raise pr and review', 0, 3.25, 0, 0, 0, 0, 0, 'PROJ-003'],
    ['MEL', 'BAU enhancement-CR', 'development and configuration', 'soanrqube issue resoltuion for Dashboard solving high issues', 0, 0, 3, 0, 0, 0, 0, 'PROJ-004'],
    ['MEL', 'BAU enhancement-CR', 'Review', 'did the unit testing and sanity of the portal and raised the pr', 0, 0, 2, 0, 0, 0, 0, 'PROJ-005'],
    ['MEL', 'BAU enhancement-CR', 'development and configuration', 'continued working on the medium and high issues', 0, 0, 2.25, 0, 0, 0, 0, 'PROJ-006'],
    ['MEL', 'BAU enhancement-CR', 'Review', 'Raised the Pr for the review', 0, 0, 0, 1.5, 0, 0, 0, 'PROJ-007'],
    ['MEL', 'BAU enhancement-CR', 'development and configuration', 'OPS>>Additional charges & Disbursement Details', 0, 0, 0, 4.25, 0, 0, 0, 'PROJ-008'],
]

for row in data:
    ws.append(row)

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 50
for col in 'EFGHIJK':
    ws.column_dimensions[col].width = 12

# Add Data Validation (Dropdowns)
dv_project = DataValidation(type='list', formula1='"MEL,Other Project"', allow_blank=True)
ws.add_data_validation(dv_project)
dv_project.add('A2:A100')

dv_sub = DataValidation(type='list', formula1='"BAU enhancement-CR,Other Sub Project"', allow_blank=True)
ws.add_data_validation(dv_sub)
dv_sub.add('B2:B100')

dv_task = DataValidation(type='list', formula1='"Meetings,development and configuration,Review,Testing,Documentation"', allow_blank=True)
ws.add_data_validation(dv_task)
dv_task.add('C2:C100')

wb.save('master_template.xlsx')
print('Master template created successfully.')
